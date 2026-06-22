"""
Migracion de CAJA Excel -> Supabase.
Uso: python scripts/migrate_caja.py "ruta/al/CAJA 2026 MARZO.xlsx"

Crea tablas caja_ingresos, caja_egresos, caja_administracion.
Ejecutar PRIMERO el SQL en supabase/caja_schema.sql
"""

import sys
import os
import re
from datetime import datetime, date
from dotenv import load_dotenv
import openpyxl
from supabase import create_client

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

load_dotenv('.env.local')
SUPABASE_URL = os.environ['NEXT_PUBLIC_SUPABASE_URL']
SUPABASE_KEY = os.environ['SUPABASE_SERVICE_ROLE_KEY']


def s(v):
    return str(v).strip() if v is not None else ''

def upper(v):
    return s(v).upper()

def to_decimal(v):
    if v is None or v == '' or v == '-':
        return 0
    try:
        return float(str(v).replace(',', '.'))
    except (ValueError, TypeError):
        return 0

def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None

def mes_from_filename(path):
    """Extrae '2026-03' del nombre del archivo"""
    name = os.path.basename(path).upper()
    meses = {
        'ENERO': '01', 'FEBRERO': '02', 'MARZO': '03', 'ABRIL': '04',
        'MAYO': '05', 'JUNIO': '06', 'JULIO': '07', 'AGOSTO': '08',
        'SEPTIEMBRE': '09', 'OCTUBRE': '10', 'NOVIEMBRE': '11', 'DICIEMBRE': '12',
    }
    m = re.search(r'20(\d\d)', name)
    year = '20' + m.group(1) if m else '2026'
    for k, v in meses.items():
        if k in name:
            return f'{year}-{v}'
    return '2026-01'


def parse_ingresos(ws, mes):
    rows = []
    all_rows = list(ws.iter_rows(min_row=8, values_only=True))
    for row in all_rows:
        if not row or len(row) < 6:
            continue
        fecha    = to_date(row[2])
        categoria = upper(row[4])
        nombre   = upper(row[5])
        detalle  = upper(row[6]) if len(row) > 6 else ''
        forma    = s(row[7])    if len(row) > 7 else ''
        monto_gs = to_decimal(row[8]) if len(row) > 8 else 0
        monto_usd= to_decimal(row[9]) if len(row) > 9 else 0
        nro_cbte = s(row[3])    if len(row) > 3 else ''

        # Saltar filas vacías
        if not fecha:
            continue
        if not categoria and not nombre and not detalle and monto_gs == 0 and monto_usd == 0:
            continue

        rows.append({
            'fecha':      fecha.isoformat(),
            'categoria':  categoria or None,
            'nombre':     nombre or None,
            'detalle':    detalle or None,
            'forma_pago': forma or None,
            'monto_gs':   monto_gs,
            'monto_usd':  monto_usd,
            'nro_cbte':   nro_cbte or None,
            'mes':        mes,
        })
    return rows


def parse_egresos(ws, mes):
    rows = []
    all_rows = list(ws.iter_rows(min_row=9, values_only=True))
    for row in all_rows:
        if not row or len(row) < 5:
            continue
        fecha      = to_date(row[0])
        nro_recibo = s(row[1])
        nro_factura= s(row[2])
        detalle    = upper(row[3])
        monto_gs   = to_decimal(row[4])
        monto_usd  = to_decimal(row[5]) if len(row) > 5 else 0

        if not fecha:
            continue
        if not detalle and monto_gs == 0 and monto_usd == 0:
            continue
        # Saltar filas de subtotal
        if nro_recibo.upper() == 'SUBTOTAL':
            continue

        rows.append({
            'fecha':       fecha.isoformat(),
            'nro_recibo':  nro_recibo or None,
            'nro_factura': nro_factura or None,
            'detalle':     detalle or None,
            'monto_gs':    monto_gs,
            'monto_usd':   monto_usd,
            'mes':         mes,
        })
    return rows


def parse_admin(ws, mes):
    rows = []
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in all_rows:
        if not row or len(row) < 3:
            continue
        fecha = to_date(row[0])
        if not fecha:
            continue
        retiro_gs  = to_decimal(row[1])
        retiro_usd = to_decimal(row[2])
        descripcion = upper(row[3]) if len(row) > 3 else ''

        if retiro_gs == 0 and retiro_usd == 0:
            continue

        rows.append({
            'fecha':       fecha.isoformat(),
            'retiro_gs':   retiro_gs,
            'retiro_usd':  retiro_usd,
            'descripcion': descripcion or None,
            'mes':         mes,
        })
    return rows


def migrate(excel_path):
    print(f'\nMigrando Caja desde: {excel_path}\n')
    mes = mes_from_filename(excel_path)
    print(f'Mes detectado: {mes}\n')

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Limpiar mes existente antes de insertar
    for tabla in ['caja_ingresos', 'caja_egresos', 'caja_administracion']:
        sb.table(tabla).delete().eq('mes', mes).execute()
        print(f'  Limpiado {tabla} para {mes}')

    # Ingresos
    ing = parse_ingresos(wb['Ingresos'], mes)
    print(f'\nIngresos encontrados: {len(ing)}')
    ok = 0
    for r in ing:
        try:
            sb.table('caja_ingresos').insert(r).execute()
            ok += 1
        except Exception as e:
            print(f'  [!] {e}')
    print(f'  Insertados: {ok}')

    # Egresos
    egr = parse_egresos(wb['Egresos'], mes)
    print(f'\nEgresos encontrados: {len(egr)}')
    ok = 0
    for r in egr:
        try:
            sb.table('caja_egresos').insert(r).execute()
            ok += 1
        except Exception as e:
            print(f'  [!] {e}')
    print(f'  Insertados: {ok}')

    # Administración
    adm = parse_admin(wb['Administración'], mes)
    print(f'\nAdministración encontrados: {len(adm)}')
    ok = 0
    for r in adm:
        try:
            sb.table('caja_administracion').insert(r).execute()
            ok += 1
        except Exception as e:
            print(f'  [!] {e}')
    print(f'  Insertados: {ok}')

    wb.close()
    print(f'\nMigracion Caja {mes} completada.\n')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Uso: python scripts/migrate_caja.py "ruta/al/archivo.xlsx"')
        sys.exit(1)
    migrate(sys.argv[1])
