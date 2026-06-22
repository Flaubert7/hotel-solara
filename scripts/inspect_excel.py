"""
Paso 1: Inspeccionar la estructura del Excel antes de migrar.
Ejecutar: python scripts/inspect_excel.py "ruta/al/archivo.xlsm"
"""

import sys
import os
import openpyxl

# Forzar UTF-8 en la salida de consola Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

def inspect(path: str):
    print(f"\nLeyendo: {path}\n")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    print(f"Hojas encontradas ({len(wb.sheetnames)}):")
    for name in wb.sheetnames:
        print(f"   - {name}")

    TARGET_SHEETS = [s for s in wb.sheetnames if any(k in s.upper() for k in [
        'RESERVAS', 'LIMP', 'INVENTARIO', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO',
        'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE', 'ENERO', 'FEBRERO',
        'MARZO', 'ABRIL', 'CONTEXTO'
    ])]

    print(f"\nHojas relevantes detectadas: {len(TARGET_SHEETS)}")

    for sheet_name in TARGET_SHEETS:
        ws = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"Hoja: {sheet_name}")
        print(f"   Primeras 5 filas:")
        rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
        for i, row in enumerate(rows, 1):
            cells = []
            for c in row[:12]:
                try:
                    cells.append(str(c)[:30] if c is not None else '-')
                except Exception:
                    cells.append('?')
            print(f"   Fila {i}: {' | '.join(cells)}")

    wb.close()
    print(f"\nInspeccion completa.\n")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python scripts/inspect_excel.py \"ruta/al/archivo.xlsm\"")
        sys.exit(1)
    inspect(sys.argv[1])
