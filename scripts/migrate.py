"""
Migracion de Excel -> Supabase para Hotel Solara.
El Excel NO se modifica. Solo se leen datos y se insertan en Supabase.

Uso:
  python scripts/migrate.py "ruta/al/archivo.xlsm"

Requiere:
  pip install openpyxl supabase python-dotenv
"""

import sys
import os
import re
from datetime import datetime, date, timedelta
from dotenv import load_dotenv
import openpyxl
from supabase import create_client

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ── Config ────────────────────────────────────────────────────────────────────
load_dotenv('.env.local')
SUPABASE_URL = os.environ['NEXT_PUBLIC_SUPABASE_URL']
SUPABASE_KEY = os.environ['SUPABASE_SERVICE_ROLE_KEY']

# Huespedes que NO deben migrarse (coliving permanente real)
SKIP_GUESTS = {'PERSONAL', '-', 'JACKEWAY', 'BROCK AWAY', 'THYM', 'THOMAS'}
# Habitaciones coliving permanente (no crear reservas para ellas)
# NOTA: 201 (WAYRA) era estadía temporal personal, SÍ debe migrarse
COLIVING_ROOMS = {'303', '304', '405', '406'}

AGENCIAS_MAP = {
    'BOOKING': 'Booking', 'AIRBNB': 'Airbnb', 'WHATSAPP': 'WhatsApp',
    'DIRECTO': 'Directo', 'INSTAGRAM': 'Instagram', 'TIKTOK': 'TikTok',
}

MESES_ES = {
    'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5, 'JUNIO': 6,
    'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9, 'OCTUBRE': 10,
    'NOVIEMBRE': 11, 'DICIEMBRE': 12, 'ENERO': 1,
    'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
    'JUL': 7, 'AGO': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12,
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def s(v) -> str:
    return str(v).strip() if v is not None else ''

def upper(v) -> str:
    return s(v).upper()

def to_date(v, year: int, month: int) -> date | None:
    if v is None:
        return None
    if isinstance(v, (datetime,)):
        return v.date()
    if isinstance(v, date):
        return v
    # Puede ser numero de dia (1-31)
    try:
        day = int(float(str(v)))
        if 1 <= day <= 31:
            try:
                return date(year, month, day)
            except ValueError:
                return None
    except (ValueError, TypeError):
        pass
    return None

def normalize_agency(v) -> str | None:
    t = upper(v)
    for k, val in AGENCIAS_MAP.items():
        if k in t:
            return val
    return s(v) or None

def to_bool(v) -> bool | None:
    t = upper(v)
    if t in ('SI', 'S', 'YES', '1', 'X', 'SÍ'):
        return True
    if t in ('NO', 'N', '0'):
        return False
    return None

def to_int(v) -> int | None:
    try:
        return int(float(str(v))) if v not in (None, '', '-') else None
    except (ValueError, TypeError):
        return None

def month_from_name(name: str) -> tuple[int, int]:
    """Retorna (mes, anio) del nombre de hoja"""
    name_up = name.upper()
    for k, v in MESES_ES.items():
        if k in name_up:
            m = re.search(r'20(\d\d)', name_up)
            year = int('20' + m.group(1)) if m else 2026
            return v, year
    return 0, 0

# ── Parsers ───────────────────────────────────────────────────────────────────

def parse_meses(wb, room_map: dict) -> list[dict]:
    """
    Lee hojas como 'MARZO 2026': matriz room x dia.
    Agrupa celdas consecutivas con mismo nombre -> reservas.
    """
    reservations = []
    month_sheets = [s for s in wb.sheetnames
                    if any(k in s.upper() for k in list(MESES_ES.keys()))
                    and '2026' in s
                    and 'RESERVAS' not in s.upper()
                    and 'LIMP' not in s.upper()]

    for sheet_name in month_sheets:
        mes, anio = month_from_name(sheet_name)
        if not mes:
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not all_rows:
            continue

        # Fila 1 = headers: [N°, HABITACION, date_col2, date_col3, ...]
        header = all_rows[0]
        # Construir mapa col_index -> date
        col_dates: dict[int, date] = {}
        for ci, cell in enumerate(header):
            if ci < 2:
                continue
            if isinstance(cell, datetime):
                col_dates[ci] = cell.date()
            elif isinstance(cell, date):
                col_dates[ci] = cell

        if not col_dates:
            print(f"  [!] {sheet_name}: sin fechas en cabecera, saltando")
            continue

        # Filas 2+ = habitaciones (saltear cualquier fila que no empiece con numero 101-406)
        for row in all_rows[1:]:
            if not row or row[0] is None:
                continue
            try:
                room_num = str(int(float(str(row[0]))))
                if not (100 <= int(room_num) <= 499):
                    continue
            except (ValueError, TypeError):
                continue

            if room_num not in room_map or room_num in COLIVING_ROOMS:
                continue

            # Agrupar consecutivas con mismo nombre
            current_guest = None
            current_start: date | None = None
            current_end: date | None = None

            for ci, d in sorted(col_dates.items()):
                cell_val = row[ci] if ci < len(row) else None
                guest = upper(cell_val) if cell_val else ''

                if guest and guest != '-' and guest not in SKIP_GUESTS:
                    if guest == current_guest:
                        current_end = d
                    else:
                        # Guardar anterior si existe
                        if current_guest and current_start:
                            reservations.append({
                                'room_num': room_num,
                                'guest_name': current_guest,
                                'check_in': current_start,
                                'check_out': current_end + timedelta(days=1) if current_end else current_start + timedelta(days=1),
                            })
                        current_guest = guest
                        current_start = d
                        current_end = d
                else:
                    if current_guest and current_start:
                        reservations.append({
                            'room_num': room_num,
                            'guest_name': current_guest,
                            'check_in': current_start,
                            'check_out': current_end + timedelta(days=1) if current_end else current_start + timedelta(days=1),
                        })
                    current_guest = None
                    current_start = None
                    current_end = None

            # Cerrar reserva abierta al final del mes
            if current_guest and current_start:
                reservations.append({
                    'room_num': room_num,
                    'guest_name': current_guest,
                    'check_in': current_start,
                    'check_out': current_end + timedelta(days=1) if current_end else current_start + timedelta(days=1),
                })

        print(f"  {sheet_name}: {sum(1 for r in reservations if r['check_in'].month == mes)} reservas acumuladas")

    return reservations


def parse_limp(wb) -> dict:
    """
    Lee hojas LIMP FEB, LIMP MAR...
    Retorna dict keyed by (upper(nombre), room_num) -> {eta, pax, breakfast, bed_type, room_type}
    """
    details: dict = {}
    limp_sheets = [s for s in wb.sheetnames if 'LIMP' in s.upper() and s.upper() != 'LIMP']

    for sheet_name in limp_sheets:
        mes, anio = month_from_name(sheet_name)
        if not mes:
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))

        # Fila 3 = "Check In | - | Nombre | HAB | Check Out | - | Noches | ETA | Pax | Des. | Obs. | Tipo"
        # Datos desde fila 5
        for row in all_rows[4:]:
            if not row or all(c is None for c in row[:6]):
                continue
            nombre = upper(row[2]) if len(row) > 2 else ''
            hab    = str(int(float(str(row[3])))) if len(row) > 3 and row[3] not in (None, '-', '') else ''
            eta    = s(row[7])  if len(row) > 7 else ''
            pax    = to_int(row[8]) if len(row) > 8 else None
            des    = to_bool(row[9]) if len(row) > 9 else None
            obs    = upper(row[10]) if len(row) > 10 else ''
            rtype  = s(row[11]) if len(row) > 11 else ''

            # Tipo de cama desde observaciones
            bed = None
            if 'MATRIMONIAL' in obs:
                bed = 'MATRIMONIAL'
            elif 'INDIVIDUAL' in obs:
                bed = 'INDIVIDUALES'

            if nombre and hab:
                key = (nombre, hab)
                details[key] = {
                    'eta': eta or None,
                    'pax': pax,
                    'breakfast': des,
                    'bed_type': bed,
                    'room_type': rtype or None,
                    'notes': obs or None,
                }
    return details


def parse_reservas_extra(wb) -> dict:
    """
    Lee hojas RESERVAS MAR, RESERVAS SEP...
    Retorna dict keyed by upper(nombre) -> {agency, modality, nights}
    """
    extra: dict = {}
    res_sheets = [s for s in wb.sheetnames if 'RESERVAS' in s.upper()]

    for sheet_name in res_sheets:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
        # header en fila 2 (index 1): - | TIPO | NOCHES | FECHA | ESTATUS | NOMBRE | AGENCIA | ...
        for row in all_rows[2:]:
            if not row or all(c is None for c in row[:6]):
                continue
            tipo   = upper(row[1]) if len(row) > 1 else ''
            noches = to_int(row[2]) if len(row) > 2 else None
            nombre = upper(row[5]) if len(row) > 5 else ''
            agencia = normalize_agency(row[6]) if len(row) > 6 else None

            if not nombre or nombre == '-':
                continue

            modality = 'COLIVING' if 'COLIVING' in tipo else 'HOTELERIA'
            if nombre not in extra:
                extra[nombre] = {'agency': agencia, 'modality': modality, 'nights': noches}
    return extra


def parse_inventario(wb) -> list[dict]:
    """Lee hoja INVENTARIO: col0=item, col2=stock inicial"""
    items = []
    inv_sheets = [s for s in wb.sheetnames if 'INVENTARIO' in s.upper()]
    if not inv_sheets:
        return items

    ws = wb[inv_sheets[0]]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    current_category = 'LIMPIEZA'
    in_data = False

    for row in all_rows:
        if not row:
            continue
        col0 = upper(row[0]) if row[0] else ''

        # Detectar inicio de seccion
        if col0 in ('LIMPIEZA', 'ACTIVOS', 'PRESTAMOS', 'ACTIVOS PARA PRESTAR'):
            current_category = 'ACTIVOS' if 'ACTIVO' in col0 or 'PREST' in col0 else 'LIMPIEZA'
            continue

        # Fila de headers (ITEM, CODIGO...)
        if col0 == 'ITEM':
            in_data = True
            continue

        if not in_data:
            continue
        if not col0 or col0 in ('-', 'TOTAL', 'SUBTOTAL'):
            continue

        stock = to_int(row[2]) if len(row) > 2 else 0
        items.append({
            'name': s(row[0]),
            'code': s(row[1]) if len(row) > 1 else None,
            'initial_stock': stock or 0,
            'category': current_category,
        })

    return items


# ── Main ──────────────────────────────────────────────────────────────────────

def deduplicate(reservations: list[dict]) -> list[dict]:
    """Elimina reservas duplicadas (misma hab + mismo huesped + misma fecha de check-in)"""
    seen = set()
    unique = []
    for r in reservations:
        key = (r['room_num'], r['guest_name'], r['check_in'])
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique


def merge_cross_month(reservations: list[dict]) -> list[dict]:
    """Une reservas del mismo huesped+hab que terminan un mes y empiezan el siguiente dia."""
    reservations = sorted(reservations, key=lambda r: (r['room_num'], r['guest_name'], r['check_in']))
    merged = []
    i = 0
    while i < len(reservations):
        r = dict(reservations[i])
        # Intentar fusionar con la siguiente
        while i + 1 < len(reservations):
            nxt = reservations[i + 1]
            if (r['room_num'] == nxt['room_num']
                    and r['guest_name'] == nxt['guest_name']
                    and r['check_out'] == nxt['check_in']):
                r['check_out'] = nxt['check_out']
                i += 1
            else:
                break
        merged.append(r)
        i += 1
    return merged


def migrate(excel_path: str):
    print(f"\nIniciando migracion desde: {excel_path}\n")

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    rooms_res = sb.table('rooms').select('id, number').execute()
    room_map = {r['number']: r['id'] for r in rooms_res.data}
    print(f"[OK] {len(room_map)} habitaciones en Supabase\n")

    # ── 1. Reservas desde MESES ───────────────────────────────────────────────
    print("Paso 1: Leyendo matrices mensuales...")
    reservations_raw = parse_meses(wb, room_map)
    reservations_raw = deduplicate(reservations_raw)
    reservations_raw = merge_cross_month(reservations_raw)
    print(f"  Total reservas unicas: {len(reservations_raw)}\n")

    # ── 2. Detalles de limpieza ───────────────────────────────────────────────
    print("Paso 2: Leyendo hojas LIMP...")
    limp_details = parse_limp(wb)
    print(f"  {len(limp_details)} registros de detalle\n")

    # ── 3. Datos extra de RESERVAS (agencia) ──────────────────────────────────
    print("Paso 3: Leyendo hojas RESERVAS (agencia)...")
    res_extra = parse_reservas_extra(wb)
    print(f"  {len(res_extra)} entradas con agencia\n")

    # ── 4. Inventario ─────────────────────────────────────────────────────────
    print("Paso 4: Leyendo INVENTARIO...")
    inv_items = parse_inventario(wb)
    print(f"  {len(inv_items)} items\n")

    wb.close()

    # ── Insertar reservas ─────────────────────────────────────────────────────
    print("Insertando reservas en Supabase...")
    res_migrated = 0
    cleaning_migrated = 0

    for r in reservations_raw:
        room_id = room_map.get(r['room_num'])
        if not room_id:
            continue

        nights = (r['check_out'] - r['check_in']).days
        guest  = r['guest_name']

        # Enriquecer con datos de limpieza
        det = limp_details.get((guest, r['room_num']), {})
        # Enriquecer con agencia
        ext = res_extra.get(guest, {})

        rec = {
            'room_id':        room_id,
            'guest_name':     guest,
            'check_in':       r['check_in'].isoformat(),
            'check_out':      r['check_out'].isoformat(),
            'nights':         nights,
            'status':         'CONFIRMADO',
            'agency':         ext.get('agency'),
            'pax':            det.get('pax'),
            'eta':            det.get('eta'),
            'bed_type':       det.get('bed_type'),
            'breakfast':      det.get('breakfast'),
            'rate_usd':       None,
            'modality':       ext.get('modality', 'HOTELERIA'),
            'notes':          det.get('notes'),
        }

        try:
            result = sb.table('reservations').insert(rec).execute()
            res_id = result.data[0]['id']
            res_migrated += 1

            # Insertar en cleaning
            cleaning_rec = {
                'reservation_id': res_id,
                'room_id':        room_id,
                'guest_name':     guest,
                'check_in_date':  r['check_in'].isoformat(),
                'check_out_date': r['check_out'].isoformat(),
                'nights':         nights,
                'eta':            det.get('eta'),
                'pax':            det.get('pax'),
                'breakfast':      det.get('breakfast'),
                'bed_type':       det.get('bed_type'),
                'room_type':      det.get('room_type'),
                'rate_usd':       None,
                'total_usd':      None,
                'is_clean':       False,
            }
            sb.table('cleaning').insert(cleaning_rec).execute()
            cleaning_migrated += 1

        except Exception as e:
            print(f"  [!] Error al insertar {guest} hab {r['room_num']} {r['check_in']}: {e}")

        if res_migrated % 20 == 0:
            print(f"  ... {res_migrated} reservas insertadas")

    # ── Insertar inventario ───────────────────────────────────────────────────
    print(f"\nInsertando inventario...")
    inv_migrated = 0
    for item in inv_items:
        if not item['name']:
            continue
        try:
            sb.table('inventory_items').insert({
                'name':          item['name'],
                'code':          item['code'] or None,
                'initial_stock': item['initial_stock'],
                'category':      item['category'],
            }).execute()
            inv_migrated += 1
        except Exception as e:
            print(f"  [!] Error item {item['name']}: {e}")

    print(f"""
{'='*50}
MIGRACION COMPLETADA

  Reservas:   {res_migrated}
  Limpieza:   {cleaning_migrated}
  Inventario: {inv_migrated}
{'='*50}
""")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Uso: python scripts/migrate.py "ruta/al/archivo.xlsm"')
        sys.exit(1)
    migrate(sys.argv[1])
