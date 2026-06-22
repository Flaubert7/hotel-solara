"""
Parche: Re-migra solo la habitacion 201 (WAYRA).
Borra los registros actuales de hab 201 y los re-crea desde el Excel.

Uso:
  python scripts/patch_hab201.py "ruta/al/archivo.xlsm"
"""

import sys
import os
from datetime import datetime, date, timedelta
from dotenv import load_dotenv
import openpyxl
from supabase import create_client

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

load_dotenv('.env.local')
SUPABASE_URL = os.environ['NEXT_PUBLIC_SUPABASE_URL']
SUPABASE_KEY = os.environ['SUPABASE_SERVICE_ROLE_KEY']

MESES_ES = {
    'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5, 'JUNIO': 6,
    'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9, 'OCTUBRE': 10,
    'NOVIEMBRE': 11, 'DICIEMBRE': 12, 'ENERO': 1,
    'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
    'JUL': 7, 'AGO': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12,
}

SKIP_GUESTS = {'PERSONAL', '-', 'JACKEWAY', 'BROCK AWAY', 'THYM', 'THOMAS'}

import re

def s(v):
    return str(v).strip() if v is not None else ''

def upper(v):
    return s(v).upper()

def month_from_name(name):
    name_up = name.upper()
    for k, v in MESES_ES.items():
        if k in name_up:
            m = re.search(r'20(\d\d)', name_up)
            year = int('20' + m.group(1)) if m else 2026
            return v, year
    return 0, 0


def leer_reservas_201(wb):
    reservations = []
    month_sheets = [s for s in wb.sheetnames
                    if any(k in s.upper() for k in list(MESES_ES.keys()))
                    and '2026' in s
                    and 'RESERVAS' not in s.upper()
                    and 'LIMP' not in s.upper()]

    for sheet_name in month_sheets:
        mes, anio = month_from_name(sheet_name)
        if not mes:
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not all_rows:
            continue

        header = all_rows[0]
        col_dates = {}
        for ci, cell in enumerate(header):
            if ci < 2:
                continue
            if isinstance(cell, datetime):
                col_dates[ci] = cell.date()
            elif isinstance(cell, date):
                col_dates[ci] = cell

        if not col_dates:
            continue

        for row in all_rows[1:]:
            if not row or row[0] is None:
                continue
            try:
                room_num = str(int(float(str(row[0]))))
                if room_num != '201':
                    continue
            except (ValueError, TypeError):
                continue

            current_guest = None
            current_start = None
            current_end = None

            for ci, d in sorted(col_dates.items()):
                cell_val = row[ci] if ci < len(row) else None
                guest = upper(cell_val) if cell_val else ''

                if guest and guest != '-' and guest not in SKIP_GUESTS:
                    if guest == current_guest:
                        current_end = d
                    else:
                        if current_guest and current_start:
                            reservations.append({
                                'guest_name': current_guest,
                                'check_in': current_start,
                                'check_out': current_end + timedelta(days=1) if current_end else current_start + timedelta(days=1),
                            })
                        current_guest = guest
                        current_start = d
                        current_end = d
                else:
                    if current_guest and current_start:
                        reservations.append({
                            'guest_name': current_guest,
                            'check_in': current_start,
                            'check_out': current_end + timedelta(days=1) if current_end else current_start + timedelta(days=1),
                        })
                    current_guest = None
                    current_start = None
                    current_end = None

            if current_guest and current_start:
                reservations.append({
                    'guest_name': current_guest,
                    'check_in': current_start,
                    'check_out': current_end + timedelta(days=1) if current_end else current_start + timedelta(days=1),
                })

        print(f"  {sheet_name}: hab 201 procesada")

    return reservations


def merge_cross_month(reservations):
    reservations = sorted(reservations, key=lambda r: (r['guest_name'], r['check_in']))
    merged = []
    i = 0
    while i < len(reservations):
        r = dict(reservations[i])
        while i + 1 < len(reservations):
            nxt = reservations[i + 1]
            if (r['guest_name'] == nxt['guest_name']
                    and r['check_out'] == nxt['check_in']):
                r['check_out'] = nxt['check_out']
                i += 1
            else:
                break
        merged.append(r)
        i += 1
    return merged


def patch(excel_path):
    print(f"\nParche hab 201 desde: {excel_path}\n")

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    room_res = sb.table('rooms').select('id, number').eq('number', '201').execute()
    if not room_res.data:
        print("[ERROR] Habitacion 201 no encontrada en la BD")
        sys.exit(1)
    room_id = room_res.data[0]['id']
    print(f"[OK] Room 201 -> id={room_id}")

    # Asegurarse que la hab 201 NO sea coliving permanente
    sb.table('rooms').update({
        'is_permanent_coliving': False,
        'coliving_guest': None
    }).eq('number', '201').execute()
    print("[OK] Room 201 marcada como NO coliving permanente")

    # Borrar reservas existentes de hab 201
    existing = sb.table('reservations').select('id').eq('room_id', room_id).execute()
    existing_ids = [r['id'] for r in existing.data]
    if existing_ids:
        sb.table('cleaning').delete().in_('reservation_id', existing_ids).execute()
        sb.table('reservations').delete().eq('room_id', room_id).execute()
        print(f"[OK] Borradas {len(existing_ids)} reservas previas de hab 201")
    else:
        print("[OK] No habia reservas previas en hab 201")

    # Leer del Excel
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    print("\nLeyendo hab 201 desde el Excel...")
    reservations = leer_reservas_201(wb)
    wb.close()

    reservations = merge_cross_month(reservations)
    print(f"\nReservas encontradas para hab 201: {len(reservations)}")
    for r in reservations:
        print(f"  {r['guest_name']}: {r['check_in']} -> {r['check_out']}")

    # Insertar
    migrated = 0
    for r in reservations:
        nights = (r['check_out'] - r['check_in']).days
        modality = 'HOTELERIA'
        # WAYRA es personal del hotel
        if r['guest_name'] in ('WAYRA',):
            modality = 'HOTELERIA'

        rec = {
            'room_id':    room_id,
            'guest_name': r['guest_name'],
            'check_in':   r['check_in'].isoformat(),
            'check_out':  r['check_out'].isoformat(),
            'nights':     nights,
            'status':     'CONFIRMADO',
            'modality':   modality,
        }
        try:
            result = sb.table('reservations').insert(rec).execute()
            res_id = result.data[0]['id']

            sb.table('cleaning').insert({
                'reservation_id': res_id,
                'room_id':        room_id,
                'guest_name':     r['guest_name'],
                'check_in_date':  r['check_in'].isoformat(),
                'check_out_date': r['check_out'].isoformat(),
                'nights':         nights,
                'is_clean':       False,
            }).execute()
            migrated += 1
            print(f"  [OK] Insertado: {r['guest_name']} {r['check_in']} -> {r['check_out']}")
        except Exception as e:
            print(f"  [!] Error: {e}")

    print(f"\nParche completado: {migrated} reservas insertadas para hab 201.\n")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Uso: python scripts/patch_hab201.py "ruta/al/archivo.xlsm"')
        sys.exit(1)
    patch(sys.argv[1])
